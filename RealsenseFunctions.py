""""

INDOOR NAVIGATION LOW COST SYSTEM FOR DRONES

Developed by: Sergi Lucas Millan
Email: sergilucasmillan@hotmail.com
Tutor: Oscar Casas Piedrafita
Center: UPC - EETAC

File description: includes functions where cameras Intel RealSense are involved

"""

import pyrealsense2 as rs
import numpy as np
import datetime
from openpyxl import *


# Configure T265 camera to stream pose data and vision from fisheye, start recording and establish connection.
def init_T265():
    pipeline_T265 = rs.pipeline()
    config_T265 = rs.config()
    config_T265.enable_device('905312110153') # Detect T265 camera by device serial
    config_T265.enable_stream(rs.stream.pose)
#    config_T265.enable_stream(rs.stream.fisheye, 1)
#    config_T265.enable_stream(rs.stream.fisheye, 2)
#    config_T265.enable_record_to_file("bagFiles/{}_T265.bag".format(datetime.datetime.now().strftime("%Y%m%d_%H%M%S"))) # Start record
    pipeline_T265.start(config_T265)
    print("LOG: T265 - Connected with camera")
#    print ("LOG: T265 - Start recording")
    return pipeline_T265

# Configure D435 camera to stream depth and color, start recording and establish connection
def init_D435():
    pipeline_D435 = rs.pipeline()
    config_D435 = rs.config()
    config_D435.enable_device('829212070982') # Detect D435 camera by device serial
    config_D435.enable_stream(rs.stream.depth)
    config_D435.enable_stream(rs.stream.color)
#    config_D435.enable_record_to_file("bagFiles/{}_D435.bag".format(datetime.datetime.now().strftime("%Y%m%d_%H%M%S"))) # Start record
    pipeline_D435.start(config_D435)
    print("LOG: D435 - Connected with camera")
#    print("LOG: D435 - Start recording")
    return pipeline_D435

# Counts number of pixels with distance below than minimum distance (security distance) and then it calculates if is an obstacle or not
def get_distance_pixels_inside_region(frames_D435, minimum_distance, x_up, y_up, x_down, y_down):
    depth_frames = frames_D435.get_depth_frame()
    list_distances = []
    j = y_up

    # Iterate matrix to get all pixel distances
    while (j < y_down):
        i = x_up
        while (i < x_down):
            pixel_distance = depth_frames.get_distance(i, j)
            if (pixel_distance < minimum_distance and pixel_distance != 0):
                list_distances.append(pixel_distance)
            i = i + 1
        j = j + 1

    # Creates histogram from all pixel distance
    hist, bins = np.histogram(list_distances, bins='auto')

    index = np.argmax(hist)

    # Observes if number of pixels below security distance is greater than fixed value (4000 in this case)
    if len(list_distances) > 4000:
        return bins[index]*100
    else:
        return 1000

    # Get distance to object (if no object is detected it returns 1 meter)
    mean_distance = bins[index]

    return mean_distance

# Save trajectory data to excel file. One file for day, and one worksheet per flight in day. (In one day it only creates
# one file but it can create many worksheets as flights in the day).
def save_to_excel(xdata,ydata,zdata):
    wb_name = "{}.xlsx".format(datetime.datetime.now().strftime("%Y%m%d")) # The file name is the day in format YYYYMMDD
    worksheet_name = str(datetime.datetime.now().strftime("%H%M%S")) # The worksheet name is the time in format hhmmss
    wb = None
    # First tries to find if there is a file created for the day. If exists, use this file. Else create a new file.
    try:
        wb = load_workbook("xlsxFiles/"+wb_name)
        print("LOG: Created new worksheet in {}" .format(wb_name))
    except:
        print("LOG: Create new .xlsx file")

    if wb is None:
        wb = Workbook()

    worksheet = wb.create_sheet(worksheet_name) # Add worksheet to excel file

    j = 1
    # Write trajectory data on the file
    while j <= len(xdata):
        worksheet.cell(row=j, column=1).value = xdata[j-1]
        worksheet.cell(row=j, column=2).value = ydata[j-1]
        worksheet.cell(row=j, column=3).value = zdata[j-1]
        j = j + 1

    wb.save("xlsxFiles/"+wb_name) # Save the file
    print("LOG: Saved trajectory data in {}" .format(wb_name))
